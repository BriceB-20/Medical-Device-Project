"""
Author: Brice Bowrey
Date: July 13th, 2020
Description: This script reads data about published research articles from Excel files and then outputs an Excel file
format of a nodes-edges table for use in network analysis. The input files are obtained from a search on PubMed's
website (examples of these files are in the repository). Each author is assigned a UUID in order to identify them in
the table. The script also counts the number of times each authors name appears in the data.
Note: It does not account for author who publish under different names and/or change their names over the years.
"""

import openpyxl
import re
import uuid

workbook = openpyxl.load_workbook(
    r"C:\Users\Briceno\Desktop\(till_2000 07062020)pubmed_medicaldevice_search 07062020.xlsx")

sheet_list = workbook.sheetnames
sheet = workbook[sheet_list[0]]

UUID_dict = {}  # {author name: author ID}
repetitions_dict = {}  # {author name: num of repetitions (e.g. number of times author's name appears)}
paper_author_list = []  # Contains lists of all author for each paper in list form
edges_dict = {}  # {author id: [matching authors id list]}

for i in range(2, sheet.max_row):
    cell = "C{}".format(str(i))

    #  Cleans cell text
    cell_author_string = str(sheet[cell].value)
    cell_author_string = re.sub("[;]", ",", cell_author_string)

    #  Creates a list of all the author names in each cell
    cell_author_list = cell_author_string.replace(".", "").split(",")
    paper_author_list.append(cell_author_list)

    for author in cell_author_list:
        index_value = cell_author_list.index(author)
        cell_author_list[index_value] = author.strip()
        if cell_author_list[index_value] == "editor" or cell_author_list[index_value] == "et al":
            del cell_author_list[index_value]

    #  Assigns each value in each list for each cell a unique ID
    for author in cell_author_list:
        try:
            x = UUID_dict[author]
            repetitions_dict[author] += 1
        except KeyError:
            UUID_dict[author] = uuid.uuid4().hex  # This can probably be improved to assure there are no duplicates
            repetitions_dict[author] = 1  # Counts the number of times an author appears in the data

# Creates edges table for network analysis in Gephi
for source_author in UUID_dict:
    edges_dict[UUID_dict[source_author]] = []
    for aut_list in paper_author_list:
        if source_author in aut_list:
            if len(aut_list) > 1:
                temp_list = aut_list.copy()
                source_author_index = temp_list.index(source_author)
                temp_list.pop(source_author_index)  # Removes the source author from the list of authors
                for target_author in temp_list:
                    edges_dict[UUID_dict[source_author]].append(UUID_dict[target_author])
            else:
                edges_dict[UUID_dict[source_author]].append(UUID_dict[source_author])

#  Data Output
data_output = openpyxl.Workbook()
ID_sheet = data_output["Sheet"]
ID_sheet.title = "ID"
edges_sheet = data_output.create_sheet("Edges")

# Labels
ID_sheet["A1"] = "Id"
ID_sheet["B1"] = "Label"
ID_sheet["C1"] = "Weight"
edges_sheet["A1"] = "Source"
edges_sheet["B1"] = "Target"

row_count = 2
for author in UUID_dict:
    a_column = "A{}".format(str(row_count))
    b_column = "B{}".format(str(row_count))
    c_column = "C{}".format(str(row_count))
    ID_sheet[a_column] = UUID_dict[author]
    ID_sheet[b_column] = author
    ID_sheet[c_column] = repetitions_dict[author]

    row_count += 1

row_count = 2
for source_author in edges_dict:
    for target_author in edges_dict[source_author]:
        a_column = "A{}".format(str(row_count))
        b_column = "B{}".format(str(row_count))
        edges_sheet[a_column] = source_author
        edges_sheet[b_column] = target_author
        row_count += 1

data_output.save(r"C:\Users\Briceno\Desktop\medicaldev-output.xlsx")
data_output.close()
