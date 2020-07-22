"""
Author: Brice Bowrey
Date: July 8th, 2020
Description: This script reads data about published research articles from Excel files and then outputs an Excel file
format of a nodes-edges table for use in network analysis. The input files are obtained from a search on PubMed's
website (examples of these files are in the repository). Each topic ("IUD", "Medical Device") is a node and an 
article about that topic creates an edge.
"""

import openpyxl
import re
import uuid

UUID_dict = {}  # {author name: author ID}
edges_dict = {}  # {topic identifier: [matching authors id list]}
cat_dict = {}  # {topic identifier: topic name string}
paper_author_list = []  # Contains lists of all author for each paper in list form


def extract_data(sheet, topic_identifier):
    for i in range(2, sheet.max_row):
        cell = "C{}".format(str(i))

        #  Cleans cell text
        cell_author_string = str(sheet[cell].value)
        cell_author_string = re.sub("[;]", ",", cell_author_string)
        cell_author_list = cell_author_string.replace(".", "").split(",")

        paper_author_list.append(cell_author_list)  # Creates a list of all the author names in each cell

        # Further cleans text
        for new_author in cell_author_list:
            index_value = cell_author_list.index(new_author)
            cell_author_list[index_value] = new_author.strip()
            if cell_author_list[index_value] == "editor" or cell_author_list[index_value] == "et al":
                del cell_author_list[index_value]

        #  Assigns each value in each list for each cell a unique ID
        for new_author in cell_author_list:
            try:
                x = UUID_dict[new_author]
                #  repetitions_dict[author] += 1
            except KeyError:
                UUID_dict[new_author] = uuid.uuid4().hex  # This can probably be improved to assure no duplicates
                #  repetitions_dict[author] = 1  # Counts the number of times an author appears in the data

        # Creates category in edges_dict and adds authors. For network analysis in Gephi.
        try:
            x = edges_dict[topic_identifier]
        except KeyError:
            edges_dict[topic_identifier] = []

        for new_author in cell_author_list:
            edges_dict[topic_identifier].append(UUID_dict[new_author])


# If this data is included in the output, the file is so large that is overwhelms Gephi
"""
Pacemakers (1)
workbook = openpyxl.load_workbook(
    r"C:\Users\Briceno\Desktop\Topic Search\csv-pacemakerT-set.xlsx")
sheet_list = workbook.sheetnames
pacemakers_sheet = workbook[sheet_list[0]]
extract_data(sheet=pacemakers_sheet, topic_identifier=1)
workbook.close()
"""

# Medical Device (2)
workbook = openpyxl.load_workbook(
    r"C:\Users\Briceno\Desktop\Topic Search\csv-medicaldev-set.xlsx")
sheet_list = workbook.sheetnames
medical_dev_sheet = workbook[sheet_list[0]]
extract_data(sheet=medical_dev_sheet, topic_identifier=2)
workbook.close()

# Joint Replacement (3)
workbook = openpyxl.load_workbook(
    r"C:\Users\Briceno\Desktop\Topic Search\csv-jointrepla-set.xlsx")
sheet_list = workbook.sheetnames
joint_rep_sheet = workbook[sheet_list[0]]
extract_data(sheet=joint_rep_sheet, topic_identifier=3)
workbook.close()

# IUD (4)
workbook = openpyxl.load_workbook(
    r"C:\Users\Briceno\Desktop\Topic Search\csv-iudTitleAb-set.xlsx")
sheet_list = workbook.sheetnames
IUD_sheet = workbook[sheet_list[0]]
extract_data(sheet=IUD_sheet, topic_identifier=4)
workbook.close()

# Artificial Lens (5)
workbook = openpyxl.load_workbook(
    r"C:\Users\Briceno\Desktop\Topic Search\csv-artificial-set.xlsx")
sheet_list = workbook.sheetnames
lens_sheet = workbook[sheet_list[0]]
extract_data(sheet=lens_sheet, topic_identifier=5)
workbook.close()


#  Data Output
data_output = openpyxl.Workbook()
ID_sheet = data_output["Sheet"]
ID_sheet.title = "ID"
# repetitions_sheet = data_output.create_sheet("Repetitions")
edges_sheet = data_output.create_sheet("Edges")

# Labels
ID_sheet["A1"] = "Id"
ID_sheet["B1"] = "Label"
edges_sheet["A1"] = "Source"
edges_sheet["B1"] = "Target"

# Produce ID sheet
count = 2
for category in edges_dict:
    a_column = "A{}".format(str(count))
    b_column = "B{}".format(str(count))
    ID_sheet[b_column] = ""  # Must fill in manually on spreadsheet
    ID_sheet[a_column] = category
    count += 1

for author in UUID_dict:
    a_column = "A{}".format(str(count))
    b_column = "B{}".format(str(count))
    ID_sheet[b_column] = author
    ID_sheet[a_column] = UUID_dict[author]
    count += 1

# Produce edges sheet
count = 2
for category in edges_dict:
    for target_author in edges_dict[category]:
        a_column = "A{}".format(str(count))
        b_column = "B{}".format(str(count))
        edges_sheet[a_column] = category
        edges_sheet[b_column] = target_author
        count += 1


data_output.save(r"C:\Users\Briceno\Desktop\(truncated)topic_search-output.xlsx")
data_output.close()
